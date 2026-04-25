# app/handlers/services/shipment_exporter.py
import os
import io
import json
import copy
import tempfile
from datetime import datetime
from decimal import Decimal
from typing import Optional, List, Tuple, Dict

import asyncio
from aiogram import Bot
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor, XDRPositiveSize2D
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from PIL import Image as PILImage


# ============================================================
# 1) CN→MSK экспорт (лист «Товары», с ФОТО)
# ============================================================

class ExcelExportService:
    """
    Экспорт посылки в Excel (лист «Товары», Китай→Москва, С ФОТО).

    - Заполнение со 2-й строки.
    - Фото берём по Telegram file_id, вписываем в реальную ячейку E{row} БЕЗ изменения
      высоты строки/ширины колонки, с небольшим внутренним отступом.
    - «цвет строго» (столбец 19 / S) — всегда 'да'.
    - «Цена доставки» (15 / O) — не трогаем.
    - «Цена (Юани)» (16 / P) — формула шаблона остаётся.
    - Возвращаем путь к временному xlsx: cargo_<ID>_<YYYYMMDD>.xlsx — отправляешь уже в своём месте.
    """

    PHOTO_COL_LETTER = "E"
    COL_INDEX = {
        "num": 1,              # A
        "title": 3,            # C
        "link": 4,             # D
        "color": 6,            # F
        "material": 7,         # G
        "cn_title": 8,         # H
        "brand": 9,            # I
        "size": 10,            # J
        "notes": 11,           # K
        "qty": 13,             # M
        "unit_price": 14,      # N
        "strict_color": 19,    # S
    }

    def __init__(
        self,
        *,
        bot: Bot,
        template_path: Optional[str] = None,
        max_concurrency: int = 8,
        photo_inner_padding_px: int = 6,
    ) -> None:
        self.bot = bot
        self.template_path = (
            template_path
            or os.getenv("CARGO_XLSX_TEMPLATE")
            or os.path.join("media", "excel", "cargo.xlsx")
        )
        self.sema = asyncio.Semaphore(int(max_concurrency))
        self.photo_inner_padding_px = int(photo_inner_padding_px)

    @staticmethod
    async def _notes_from_extra(extra: object) -> str:
        KEYS = ("note", "comment", "备注", "примечание", "notes")
        if extra is None:
            return ""
        if isinstance(extra, dict):
            for k in KEYS:
                v = extra.get(k)
                if v:
                    return str(v)
            return ""
        if isinstance(extra, str):
            try:
                data = json.loads(extra)
                if isinstance(data, dict):
                    for k in KEYS:
                        v = data.get(k)
                        if v:
                            return str(v)
            except Exception:
                pass
        return ""

    @staticmethod
    async def _extra_field(extra: object, keys: Tuple[str, ...]) -> str:
        if extra is None:
            return ""
        if isinstance(extra, dict):
            for k in keys:
                v = extra.get(k)
                if v not in (None, ""):
                    return str(v)
            return ""
        if isinstance(extra, str):
            try:
                data = json.loads(extra)
                if isinstance(data, dict):
                    for k in keys:
                        v = data.get(k)
                        if v not in (None, ""):
                            return str(v)
            except Exception:
                pass
        return ""

    @staticmethod
    async def _points_to_pixels(points: float) -> float:
        return points * 4.0 / 3.0

    @staticmethod
    async def _col_width_chars_to_pixels(chars_width: float) -> float:
        return chars_width * 7.0

    @staticmethod
    async def _worksheet_default_row_height_pt(ws: Worksheet) -> float:
        return float(getattr(ws.sheet_format, "defaultRowHeight", 15.0) or 15.0)

    @staticmethod
    async def _worksheet_default_col_width_chars(ws: Worksheet) -> float:
        return float(getattr(ws.sheet_format, "defaultColWidth", 8.43) or 8.43)

    async def _get_cell_box_pixels(self, ws: Worksheet, row: int, col_letter: str) -> Tuple[int, int]:
        col_dim = ws.column_dimensions.get(col_letter)
        if col_dim and col_dim.width:
            col_chars = float(col_dim.width)
        else:
            col_chars = await self._worksheet_default_col_width_chars(ws)
        width_px = await self._col_width_chars_to_pixels(col_chars)
        row_dim = ws.row_dimensions.get(row)
        if row_dim and row_dim.height:
            row_pt = float(row_dim.height)
        else:
            row_pt = await self._worksheet_default_row_height_pt(ws)
        height_px = await self._points_to_pixels(row_pt)
        return int(width_px), int(height_px)

    async def _download_photo_bytes(self, file_id: Optional[str]) -> Optional[bytes]:
        if not file_id:
            return None
        async with self.sema:
            try:
                tg_file = await self.bot.get_file(file_id)
                buf = io.BytesIO()
                await self.bot.download(tg_file, destination=buf)
                return buf.getvalue()
            except Exception:
                return None

    async def _process_image_to_png_fit_box(
        self,
        raw: bytes,
        box_w_px: int,
        box_h_px: int,
        padding_px: int,
    ) -> Optional[Tuple[bytes, int, int]]:
        def _work() -> Optional[Tuple[bytes, int, int]]:
            try:
                img = PILImage.open(io.BytesIO(raw)).convert("RGB")
                w, h = img.size
                tgt_w = max(1, box_w_px - 2 * padding_px)
                tgt_h = max(1, box_h_px - 2 * padding_px)
                scale = min(tgt_w / max(1, w), tgt_h / max(1, h))
                new_w = max(1, int(w * scale))
                new_h = max(1, int(h * scale))
                if scale < 1.0:
                    img = img.resize((new_w, new_h), PILImage.LANCZOS)
                out = io.BytesIO()
                img.save(out, format="PNG", optimize=True)
                return out.getvalue(), new_w, new_h
            except Exception:
                return None
        return await asyncio.to_thread(_work)

    async def generate_goods_sheet(self, *, cargo_service, cargo_id: int) -> str:
        items: List[dict] = await cargo_service.items.list_by_cargo(cargo_id=cargo_id)
        file_ids = [it.get("photo_file_id") for it in items]
        photos_raw = await asyncio.gather(*[self._download_photo_bytes(fid) for fid in file_ids])

        if not os.path.exists(self.template_path):
            raise FileNotFoundError(f"Excel-шаблон не найден: {self.template_path}")
        wb = load_workbook(self.template_path)
        ws = wb["Товары"] if "Товары" in wb.sheetnames else wb.active

        row = 2
        cell = ws.cell
        for i, it in enumerate(items, 1):
            title = it.get("title") or "Без названия"
            link = it.get("source_url") or ""
            color = it.get("color") or ""
            size = it.get("size") or ""
            qty = int(it.get("quantity") or 0)
            price = it.get("price") or 0
            notes = await self._notes_from_extra(it.get("extra"))

            extra = it.get("extra")
            material = await self._extra_field(extra, ("material", "материал", "Материал", "材质"))
            cn_title = await self._extra_field(extra, (
                "cn_title", "chinese_name", "chinese_title", "中文品名", "中文名称", "名称"
            ))
            brand = await self._extra_field(extra, ("brand", "бренд", "Бренд", "品牌"))

            cell(row=row, column=self.COL_INDEX["num"],        value=i)
            cell(row=row, column=self.COL_INDEX["title"],      value=str(title))
            cell(row=row, column=self.COL_INDEX["link"],       value=str(link))
            cell(row=row, column=self.COL_INDEX["color"],      value=str(color))
            cell(row=row, column=self.COL_INDEX["material"],   value=str(material))
            cell(row=row, column=self.COL_INDEX["cn_title"],   value=str(cn_title))
            cell(row=row, column=self.COL_INDEX["brand"],      value=str(brand))
            cell(row=row, column=self.COL_INDEX["size"],       value=str(size))
            cell(row=row, column=self.COL_INDEX["notes"],      value=str(notes))
            cell(row=row, column=self.COL_INDEX["qty"],        value=qty)
            try:
                unit_price = float(Decimal(str(price)))
            except Exception:
                unit_price = 0.0
            cell(row=row, column=self.COL_INDEX["unit_price"], value=unit_price)
            cell(row=row, column=self.COL_INDEX["strict_color"], value="да")

            raw = photos_raw[i - 1]
            if raw:
                box_w_px, box_h_px = await self._get_cell_box_pixels(ws, row=row, col_letter=self.PHOTO_COL_LETTER)
                processed = await self._process_image_to_png_fit_box(
                    raw=raw,
                    box_w_px=box_w_px,
                    box_h_px=box_h_px,
                    padding_px=self.photo_inner_padding_px,
                )
                if processed:
                    png_bytes, w_px, h_px = processed
                    bio = io.BytesIO(png_bytes)
                    xl_img = XLImage(bio)
                    xl_img.width = w_px
                    xl_img.height = h_px

                    offset_x = max(0, (box_w_px - w_px) // 2)
                    offset_y = max(0, (box_h_px - h_px) // 2)

                    col_idx_1based = column_index_from_string(self.PHOTO_COL_LETTER)
                    marker = AnchorMarker(
                        col=col_idx_1based - 1,
                        colOff=offset_x * 9525,
                        row=row - 1,
                        rowOff=offset_y * 9525,
                    )
                    ext = XDRPositiveSize2D(cx=w_px * 9525, cy=h_px * 9525)
                    anchor = OneCellAnchor(_from=marker, ext=ext)
                    ws.add_image(xl_img, anchor)

            row += 1

        today = datetime.now().strftime("%Y%m%d")
        tmpdir = tempfile.gettempdir()
        file_path = os.path.join(tmpdir, f"cargo_{cargo_id}_{today}.xlsx")
        wb.save(file_path)
        wb.close()
        return file_path


# ============================================================
# 2) Текстовый бланк «Садовод» (без фото)
# ============================================================

class ExcelTextFormExportService:
    """
    Экспорт «текстового» бланка (второй шаблон, без фото) из БД.
    """

    def __init__(
        self,
        *,
        template_path: Optional[str] = None,
        sheet_name: Optional[str] = None,
        start_row: int = 5,
        end_row: int = 33,
        total_label: str = "Всего",
        yuan_to_rub: Optional[Decimal] = None,
    ) -> None:
        self.template_path = (
            template_path
            or os.getenv("CARGO_XLSX_TEMPLATE")
            or os.path.join("media", "excel", "sadovod.xlsx")
        )
        if not self.template_path:
            raise ValueError("Не указан template_path и не задан ENV TEXT_FORM_XLSX_TEMPLATE")
        self.sheet_name = sheet_name
        self.start_row = int(start_row)
        self.end_row = int(end_row)
        self.total_label = total_label
        env_rate = os.getenv("YUAN_TO_RUB") or "15"
        self.yuan_to_rub = Decimal(str(yuan_to_rub)) if yuan_to_rub is not None else Decimal(env_rate)
        self._total_row_cache: Dict[int, dict] = {}

        self.column_mapping = {
            "title":       list(range(3,  9)),
            "unit":        list(range(12, 14)),
            "qty":         list(range(18, 22)),
            "unit_price":  list(range(22, 27)),
        }

    async def _copy_row_format(self, sheet, source_row: int, target_row: int) -> None:
        for col in range(1, sheet.max_column + 1):
            src = sheet.cell(row=source_row, column=col)
            tgt = sheet.cell(row=target_row, column=col)
            if getattr(src, "has_style", False):
                tgt.font = copy.copy(src.font)
                tgt.border = copy.copy(src.border)
                tgt.fill = copy.copy(src.fill)
                tgt.number_format = copy.copy(src.number_format)
                tgt.protection = copy.copy(src.protection)
                tgt.alignment = copy.copy(src.alignment)

    async def _remove_merged_ranges_on_row(self, sheet, row: int) -> None:
        to_remove = [rng for rng in sheet.merged_cells.ranges if rng.min_row == row and rng.max_row == row]
        for rng in to_remove:
            sheet.unmerge_cells(str(rng))

    async def _move_merged_cells(self, sheet, source_row: int, target_row: int) -> None:
        to_copy = [rng for rng in sheet.merged_cells.ranges if rng.min_row == source_row and rng.max_row == source_row]
        for rng in to_copy:
            coord = f"{get_column_letter(rng.min_col)}{target_row}:{get_column_letter(rng.max_col)}{target_row}"
            sheet.merge_cells(coord)

    async def _find_total_row(self, sheet) -> Optional[int]:
        for row in sheet.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and self.total_label in v:
                    return cell.row
        return None

    async def _cache_total_row(self, sheet, row: int) -> None:
        self._total_row_cache.clear()
        for col in range(1, sheet.max_column + 1):
            c = sheet.cell(row=row, column=col)
            self._total_row_cache[col] = {
                "value": c.value,
                "font": copy.copy(c.font),
                "border": copy.copy(c.border),
                "fill": copy.copy(c.fill),
                "alignment": copy.copy(c.alignment),
                "number_format": c.number_format,
            }
        self._total_row_cache["merges"] = [
            (rng.min_col, rng.max_col)
            for rng in sheet.merged_cells.ranges
            if rng.min_row == row and rng.max_row == row
        ]

    async def _restore_total_row(self, sheet, target_row: int) -> None:
        sheet.insert_rows(target_row)
        for col, data in self._total_row_cache.items():
            if col == "merges":
                continue
            cell = sheet.cell(row=target_row, column=int(col))
            cell.value = data["value"]
            cell.font = copy.copy(data["font"])
            cell.border = copy.copy(data["border"])
            cell.fill = copy.copy(data["fill"])
            cell.alignment = copy.copy(data["alignment"])
            cell.number_format = data["number_format"]
        for min_col, max_col in self._total_row_cache.get("merges", []):
            coord = f"{get_column_letter(min_col)}{target_row}:{get_column_letter(max_col)}{target_row}"
            sheet.merge_cells(coord)

    async def _safe_set_cell(self, sheet, row: int, col: int, value) -> None:
        cell = sheet.cell(row=row, column=col)
        if type(cell).__name__ == "MergedCell":
            for merged in sheet.merged_cells.ranges:
                if cell.coordinate in merged:
                    top_left = sheet.cell(row=merged.min_row, column=merged.min_col)
                    top_left.value = value
                    return
        else:
            cell.value = value

    async def _load_rows_from_db(self, cargo_service, cargo_id: int) -> List[dict]:
        items = await cargo_service.items.list_by_cargo(cargo_id=cargo_id)
        rows: List[dict] = []
        for it in items:
            title = it.get("title") or "Без названия"
            qty = int(it.get("quantity") or 0)
            cny = Decimal(str(it.get("price") or 0))
            rub = (cny * self.yuan_to_rub).quantize(Decimal("0.01"))
            rows.append({"title": title, "qty": qty, "unit": "штука, шт.", "unit_price_rub": rub})
        return rows

    async def generate_text_form(self, *, cargo_service, cargo_id: int) -> str:
        if not os.path.exists(self.template_path):
            raise FileNotFoundError(f"Excel-шаблон не найден: {self.template_path}")

        rows = await self._load_rows_from_db(cargo_service, cargo_id)
        rows_count = len(rows)

        today = datetime.now().strftime("%Y%m%d")
        tmpdir = tempfile.gettempdir()
        file_path = os.path.join(tmpdir, f"cargo_{cargo_id}_{today}_form.xlsx")

        wb = load_workbook(self.template_path)
        ws = wb[self.sheet_name] if (self.sheet_name and self.sheet_name in wb.sheetnames) else wb.active

        start_row = self.start_row
        end_row = self.end_row
        template_row = start_row
        max_existing = end_row - start_row + 1

        total_row_original = await self._find_total_row(ws)
        if total_row_original is None:
            wb.close()
            raise RuntimeError(f"Не найдена строка с текстом '{self.total_label}'")

        move_threshold = int(getattr(self, "move_total_threshold", 29))
        move_total = rows_count > move_threshold

        if move_total:
            await self._cache_total_row(ws, total_row_original)
            await self._remove_merged_ranges_on_row(ws, total_row_original)
            ws.delete_rows(total_row_original)

        if move_total:
            capacity = max_existing
        else:
            capacity = min(max_existing, (total_row_original - start_row))

        rows_written = 0
        write_count = min(rows_count, capacity)
        for i in range(write_count):
            row_idx = start_row + i
            r = rows[i]
            await self._safe_set_cell(ws, row_idx, 2, i + 1)
            for col in self.column_mapping["title"]:
                await self._safe_set_cell(ws, row_idx, col, r["title"])
            for col in self.column_mapping["unit"]:
                await self._safe_set_cell(ws, row_idx, col, r["unit"])
            for col in self.column_mapping["qty"]:
                await self._safe_set_cell(ws, row_idx, col, r["qty"])
            for col in self.column_mapping["unit_price"]:
                await self._safe_set_cell(ws, row_idx, col, r["unit_price_rub"])

        rows_written = write_count

        if move_total and rows_count > capacity:
            for i in range(capacity, rows_count):
                row_idx = start_row + i
                ws.insert_rows(row_idx)
                await self._copy_row_format(ws, template_row, row_idx)
                await self._move_merged_cells(ws, template_row, row_idx)

                r = rows[i]
                await self._safe_set_cell(ws, row_idx, 2, i + 1)
                for col in self.column_mapping["title"]:
                    await self._safe_set_cell(ws, row_idx, col, r["title"])
                for col in self.column_mapping["unit"]:
                    await self._safe_set_cell(ws, row_idx, col, r["unit"])
                for col in self.column_mapping["qty"]:
                    await self._safe_set_cell(ws, row_idx, col, r["qty"])
                for col in self.column_mapping["unit_price"]:
                    await self._safe_set_cell(ws, row_idx, col, r["unit_price_rub"])

            rows_written = rows_count

        if move_total:
            final_row = start_row + rows_written
            await self._restore_total_row(ws, final_row)

        wb.save(file_path)
        wb.close()
        return file_path


# ============================================================
# 3) ТК Экспедиция (авто-выбор листа по кол-ву товаров)
# ============================================================

# Правила выбора листа из шаблона (по документу Образец):
# Лист 2 → товаров 1-31
# Лист 3 → товаров 32-83
# Лист 4 → товаров 84-134
# Лист 5 → товаров 135-185
_EXPEDITION_SHEET_RANGES: List[Tuple[int, int, str]] = [
    (1,   31,  "1-31"),
    (32,  83,  "32-83"),
    (84,  134, "84-134"),
    (135, 185, "135-185"),
]


class ExcelExpeditionExportService:
    """
    Экспорт ТК «Экспедиция» (сопроводительное письмо).

    Бот автоматически выбирает нужный лист шаблона по количеству товаров:
      1–31   → лист «1-31»
      32–83  → лист «32-83»
      84–134 → лист «84-134»
      135–185 → лист «135-185»

    Все единицы измерения — «шт.».
    Столбцы (C=Наименование, D=Кол-во, E=Ед. изм., F=Стоимость):
      «Наименование товара» — название из БД
      «Кол-во единиц»       — quantity из БД
      «Единица измерения»   — всегда «шт.»
      «Общая стоимость»     — price * quantity * yuan_to_rub (бел. рубли)

    Первая строка данных определяется по первой пустой ячейке в колонке
    номера (#), начиная с HEADER_ROW_OFFSET.
    """

    # Заголовок таблицы всегда занимает несколько строк.
    # Ищем первую строку с числом «1» в первом столбце — это строка данных.
    _DATA_SEARCH_MAX_ROW = 30  # не более 30 строк шапки

    # Колонки данных (1-based). Берём из образца файла:
    # B(2) = №, C(3) = Наименование, D(4) = Кол-во, E(5) = Ед.изм., F(6) = Стоимость
    _COL_NUM   = 2
    _COL_TITLE = 3
    _COL_QTY   = 4
    _COL_UNIT  = 5
    _COL_PRICE = 6

    def __init__(
        self,
        *,
        template_path: Optional[str] = None,
        yuan_to_rub: Optional[Decimal] = None,
    ) -> None:
        """
        :param template_path: путь к Excel-шаблону ТК Экспедиция;
                              по умолчанию: env EXPEDITION_XLSX_TEMPLATE
                              или media/excel/expedition.xlsx
        :param yuan_to_rub:   курс CNY→BYR; по умолчанию: env YUAN_TO_RUB или 15
        """
        self.template_path = (
            template_path
            or os.getenv("EXPEDITION_XLSX_TEMPLATE")
            or os.path.join("media", "excel", "expedition.xlsx")
        )
        env_rate = os.getenv("YUAN_TO_RUB") or "15"
        self.yuan_to_rub = (
            Decimal(str(yuan_to_rub))
            if yuan_to_rub is not None
            else Decimal(env_rate)
        )

    # ------------------------------------------------------------------
    # helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _select_sheet_name(items_count: int) -> str:
        """
        Возвращает имя листа по количеству товаров.
        Если > 185 — возвращает последний лист (185 строк).
        """
        for lo, hi, name in _EXPEDITION_SHEET_RANGES:
            if lo <= items_count <= hi:
                return name
        # items_count == 0 или > 185
        if items_count == 0:
            return _EXPEDITION_SHEET_RANGES[0][2]  # «1-31»
        return _EXPEDITION_SHEET_RANGES[-1][2]      # «135-185»

    @staticmethod
    def _find_data_start_row(ws, max_search: int = 30) -> int:
        """
        Ищет первую строку, где в колонке _COL_NUM (B) стоит число 1
        (начало таблицы данных). Fallback — возвращает max_search + 1.
        """
        for r in range(1, max_search + 1):
            val = ws.cell(row=r, column=ExcelExpeditionExportService._COL_NUM).value
            if val == 1 or val == "1":
                return r
        return max_search + 1

    @staticmethod
    async def _safe_write(ws, row: int, col: int, value) -> None:
        """Пишет значение, обходя MergedCell."""
        cell = ws.cell(row=row, column=col)
        if type(cell).__name__ == "MergedCell":
            for merged in ws.merged_cells.ranges:
                if cell.coordinate in merged:
                    ws.cell(row=merged.min_row, column=merged.min_col).value = value
                    return
        else:
            cell.value = value

    async def _load_items(self, cargo_service, cargo_id: int) -> List[dict]:
        """Загружает товары из БД."""
        items = await cargo_service.items.list_by_cargo(cargo_id=cargo_id)
        result: List[dict] = []
        for it in items:
            title = it.get("title") or "Без названия"
            qty   = int(it.get("quantity") or 1)
            cny   = Decimal(str(it.get("price") or 0))
            total_rub = (cny * self.yuan_to_rub * qty).quantize(Decimal("0.01"))
            result.append({
                "title":     title,
                "qty":       qty,
                "unit":      "шт.",
                "total_rub": total_rub,
            })
        return result

    # ------------------------------------------------------------------
    # публичный метод
    # ------------------------------------------------------------------

    async def generate(
        self,
        *,
        cargo_service,
        cargo_id: int,
    ) -> str:
        """
        Заполняет шаблон ТК Экспедиция и возвращает путь к временному xlsx.

        :param cargo_service: CargoService
        :param cargo_id:      ID посылки
        :return:              /tmp/expedition_<cargo_id>_<YYYYMMDD>.xlsx
        """
        if not os.path.exists(self.template_path):
            raise FileNotFoundError(
                f"Excel-шаблон ТК Экспедиция не найден: {self.template_path}"
            )

        items = await self._load_items(cargo_service, cargo_id)
        count = len(items)
        sheet_name = self._select_sheet_name(count)

        wb = load_workbook(self.template_path)

        if sheet_name not in wb.sheetnames:
            wb.close()
            raise RuntimeError(
                f"Лист '{sheet_name}' не найден в шаблоне. "
                f"Доступные листы: {wb.sheetnames}"
            )

        ws = wb[sheet_name]
        data_start = self._find_data_start_row(ws, self._DATA_SEARCH_MAX_ROW)

        for idx, it in enumerate(items):
            row = data_start + idx
            await self._safe_write(ws, row, self._COL_NUM,   idx + 1)
            await self._safe_write(ws, row, self._COL_TITLE, it["title"])
            await self._safe_write(ws, row, self._COL_QTY,   it["qty"])
            await self._safe_write(ws, row, self._COL_UNIT,  it["unit"])
            await self._safe_write(ws, row, self._COL_PRICE, float(it["total_rub"]))

        today = datetime.now().strftime("%Y%m%d")
        tmpdir = tempfile.gettempdir()
        out_path = os.path.join(
            tmpdir, f"expedition_{cargo_id}_{today}.xlsx"
        )
        wb.save(out_path)
        wb.close()
        return out_path


# ============================================================
# 4) ФАСАДЫ
# ============================================================

async def export_cn_msk_goods(
    *,
    bot: Bot,
    cargo_service,
    cargo_id: int,
    template_path: Optional[str] = None,
    max_concurrency: int = 8,
    photo_inner_padding_px: int = 6,
) -> str:
    """Фасад для CN→MSK (с фото)."""
    svc = ExcelExportService(
        bot=bot,
        template_path=template_path,
        max_concurrency=max_concurrency,
        photo_inner_padding_px=photo_inner_padding_px,
    )
    return await svc.generate_goods_sheet(
        cargo_service=cargo_service, cargo_id=cargo_id
    )


async def export_text_form(
    *,
    cargo_service,
    cargo_id: int,
    template_path: Optional[str] = None,
    sheet_name: Optional[str] = None,
    start_row: int = 5,
    end_row: int = 33,
    total_label: str = "Всего",
    yuan_to_rub: Optional[Decimal] = None,
) -> str:
    """Фасад для текстового бланка Садовод (без фото)."""
    svc = ExcelTextFormExportService(
        template_path=template_path,
        sheet_name=sheet_name,
        start_row=start_row,
        end_row=end_row,
        total_label=total_label,
        yuan_to_rub=yuan_to_rub,
    )
    return await svc.generate_text_form(
        cargo_service=cargo_service, cargo_id=cargo_id
    )


async def export_expedition(
    *,
    cargo_service,
    cargo_id: int,
    template_path: Optional[str] = None,
    yuan_to_rub: Optional[Decimal] = None,
) -> str:
    """Фасад для ТК Экспедиция (авто-выбор листа по кол-ву товаров)."""
    svc = ExcelExpeditionExportService(
        template_path=template_path,
        yuan_to_rub=yuan_to_rub,
    )
    return await svc.generate(
        cargo_service=cargo_service, cargo_id=cargo_id
    )
